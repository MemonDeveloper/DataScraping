import tkinter as tk
import threading
import json
import shutil
import datetime
import pandas as pd
import undetected_chromedriver as uc
import requests
import concurrent.futures
import os
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from selenium.common.exceptions import NoSuchElementException
from tkinter import messagebox
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait as wait
from selenium.webdriver.common.keys import Keys

# Global flag to control whether the task is running
running = True
memon_kill = ""

def clean_Account(x):
    if pd.isna(x):
        return None
    try:
        if isinstance(x, float) and x.is_integer():
            return str(int(x))
        return str(x)
    except:
        return str(x)

def on_closing():
    global running
    running = False  # Stop the task
    root.quit()  # Close the Tkinter window

# Function to safely update the TextBox
def update_username(text_widget, message):
    # Use the after() method to safely update the widget in the main thread
    text_widget.after(0, lambda: text_widget.config(state=tk.NORMAL))  # Ensure widget is in NORMAL state
    text_widget.after(0, lambda: text_widget.insert(tk.END, message + "\n"))
    text_widget.after(0, lambda: text_widget.see(tk.END))
    text_widget.after(0, lambda: text_widget.config(state=tk.DISABLED))  # Disable the widget after update

def on_exit():
    global running
    running = False  # Stop the task
    root.quit()  # This ensures the program stops correctly

# Function to stop the thread
def run_stop_script():
    global running
    running = False  # Stop the task
    memon_kill = "killthread"

def run_bahsegel1200_script():
    global running
    # Clear the textbox when button is clicked
    output_text.config(state=tk.NORMAL)
    output_text.delete("1.0", tk.END)
    output_text.config(state=tk.DISABLED)

    def task():
        original_file_path = "bahsegel1200.xlsx"
        # Check if the file exists
        if os.path.exists(original_file_path):
            try:
                workbook = load_workbook(original_file_path, read_only=True)
                sheet_names = workbook.sheetnames

                if "bahsegel1200 List" in sheet_names:
                    start_time = datetime.datetime.now().time()
                    update_username(output_text, f"Please wait while Chrome is starting up. (bahsegel1200) {start_time}")
                    # Setup undetected Chrome
                    options = uc.ChromeOptions()
                    # Set up to block unnecessary background requests (external resources)
                    # options.add_argument("--headless")  # Run in headless mode (no GUI)
                    # options.add_argument("--headless=new")
                    # ooptions.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36")
                    options.add_argument("--disable-blink-features=AutomationControlled")  # Disable automation flag
                    options.add_argument("--disable-extensions")  # Disable extensions
                    options.add_argument("--disable-background-networking")  # Disable background networking
                    options.add_argument("--disable-software-rasterizer")  # Disable software rasterizer
                    options.add_argument("--disable-gpu")
                    options.add_argument("--no-sandbox")
                    options.add_argument("--log-level=3")
                    options.add_argument("--start-minimized")

                    driver = uc.Chrome(options=options)  # Adjust version to match your installed Chrome
                    update_username(output_text, f"Google Chrome has successfully launched.")
                    # Setup Excel files
                    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                    backup_file_path = f"bahsegel1200 - Result {timestamp}.xlsx"
                    shutil.copy(original_file_path, backup_file_path)

                    # Read Accounts
                    all_sheets = pd.read_excel(original_file_path, sheet_name=["bahsegel1200 List"])
                    df_Account = all_sheets["bahsegel1200 List"]
                    Account_list = df_Account.iloc[:, 0].apply(clean_Account).tolist()
                    password_list = df_Account.iloc[:, 1].apply(clean_Account).tolist()

                    product_list = []
                    success_count = 0
                    failure_count = 0
                    remaining = len(Account_list)

                    for Account, password in zip(Account_list, password_list):
                        username = ""
                        firstname = ""
                        lastname = ""
                        address = ""
                        postalcode = ""
                        brithdate = ""
                        city = ""
                        country = ""
                        email = ""
                        mobile_phone = ""
                        totalbalance = ""
                        bonusbalance = ""
                        realbalance = ""
                        tokenbalance = ""

                        if Account is None:
                            product_list.append(["Account Not Found"] + [""] * 13)
                            update_username(None, f"❌ Account {Account} Remaining {remaining}: Not Found")
                            failure_count += 1
                            remaining -= 1
                        else:
                            url = "https://www.bahsegel1200.com/#login"
                            try:
                                driver.get(url)
                                wait_driver = wait(driver, 10)
                            except Exception as e:
                                update_username(None, f"❌ Failed to load login page for {Account}: {str(e)}")
                                failure_count += 1
                                remaining -= 1
                                continue
                            try:
                                access = wait_driver.until(EC.presence_of_element_located(
                                    (By.XPATH, '//h2[contains(text(), "Sorry, but Bahsegel is not available in your country")]')
                                ))
                                access = "Sorry, but Bahsegel is not available in your country"
                            except Exception as e:
                                access = ""

                            if access == "Sorry, but Bahsegel is not available in your country":
                                product_list.append(["Sorry, but Bahsegel is not available in your country"] + [""] * 13)
                                update_username(None, f"❌ Account {Account} Remaining {remaining}: Not Found")
                                failure_count += 1
                                remaining -= 1
                            else:
                                try:
                                    wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="login-btn"]')))
                                    # Input fields
                                    username_field = wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="username-input"]')))
                                    # password_field = wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="password-input"]')))
                                    username_field.send_keys(Account + Keys.TAB + password)
                                    # Click login
                                    login_button = wait_driver.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-testid="login-btn"]')))
                                    login_button.click()
                                    
                                    try:
                                        wait_driver.until(EC.presence_of_element_located((By.CLASS_NAME, "styles__TrueplayProposalDesc-sc-1j8w4od-22")))
                                    except Exception as e:
                                        wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#user-info--as-dropdown > div")))

                                    driver.get("https://www.bahsegel1200.com/en/profile/personal-info")

                                    # Get all balance blocks including Token Balance
                                    logout_button = wait_driver.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.btn.btn-xs.btn-transparent-bordered.btn-block')))
                                    balance_blocks = wait_driver.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".account-info__balance-block")))

                                    balances = {}
                                    for block in balance_blocks:
                                        label_span = block.find_element(By.CLASS_NAME, "account-info__balance-block-text")
                                        # Remove any nested labels like "New"
                                        label = label_span.find_elements(By.XPATH, ".//span[@class='status-label__text']")
                                        for nested in label:
                                            driver.execute_script("arguments[0].remove();", nested)

                                        # Get cleaned label and value
                                        label_clean = label_span.text.strip()
                                        value = block.find_element(By.CLASS_NAME, "account-info__balance-block-value").text.strip()
                                        balances[label_clean] = value

                                    totalbalance = balances.get('Total balance', 'N/A')
                                    bonusbalance = balances.get('Bonus balance', 'N/A')
                                    realbalance = balances.get('Real balance', 'N/A')
                                    tokenbalance = balances.get('Token Balance', 'N/A')

                                    username_element = wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="username-input"]')))
                                    username = username_element.get_attribute("value")
                                    firstname_element = wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="first-name-input"]')))
                                    firstname = firstname_element.get_attribute("value")
                                    lastname_element = wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="last-name-input"]')))
                                    lastname = lastname_element.get_attribute("value")
                                    address_element = wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="address-input"]')))
                                    address = address_element.get_attribute("value")
                                    postal_code_element = wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="postal-code-input"]')))
                                    postalcode = postal_code_element.get_attribute("value")
                                    brithdate_element = wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[name="birth_date"]')))
                                    brithdate = brithdate_element.get_attribute("value")
                                    city_element = wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="city-input"]')))
                                    city = city_element.get_attribute("value")
                                    
                                    # Extract visible country name (e.g., "Turkey")
                                    country_label_element = wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="country-dropdown"] label')))
                                    country = country_label_element.text.strip()

                                    # # Extract country code (e.g., "TR")
                                    # country_input_element = wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="country-dropdown"] input[name="country"]')))
                                    # country = country_input_element.get_attribute("value")

                                    email_element = wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="email-input"]')))
                                    email = email_element.get_attribute("value")
                                    mobile_phone_element = wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="mobile-phone-input"]')))
                                    mobile_phone = mobile_phone_element.get_attribute("value")

                                    # Log out
                                    logout_button = wait_driver.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.btn.btn-xs.btn-transparent-bordered.btn-block')))
                                    logout_button.click()
                                    wait_driver.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-testid="login-btn"]')))

                                    product_list.append([
                                        username,
                                        firstname,
                                        lastname,
                                        address,
                                        postalcode,
                                        brithdate,
                                        city,
                                        country,
                                        email,
                                        mobile_phone,
                                        totalbalance,
                                        bonusbalance,
                                        realbalance,
                                        tokenbalance])
                                    remaining -=1
                                    success_count += 1
                                    update_username(output_text, f"✅ Account {Account} Remaining {remaining}: Success")

                                except Exception as e:
                                    update_username(None, f"❌ Login failed for {Account}: {str(e)}")
                                    failure_count += 1

                    # Quit browser
                    driver.quit()

                    # Prepare Excel output
                    book = load_workbook(backup_file_path)
                    df_existing = pd.read_excel(backup_file_path, sheet_name="bahsegel1200 List", usecols="A:B")

                    df_result = pd.DataFrame(product_list, columns=[
                        "username", "firstname", "lastname", "address", "postalcode", "brithdate", "city",
                        "country", "email", "mobile_phone", "totalbalance", "bonusbalance", "realbalance", "tokenbalance"
                    ])

                    df_combined = pd.concat([df_existing, df_result], axis=1)

                    # ========== STEP 8: SAVE THE COMBINED DATA BACK TO THE BACKUP FILE ==========

                    if 'Result' in book.sheetnames:
                        del book['Result']
                        book.save(backup_file_path)  # Save after deleting to avoid error later

                    # Write new sheet
                    with pd.ExcelWriter(backup_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_combined.to_excel(writer, sheet_name="Result", index=False)

                    # ========== STEP 9: FORMAT THE SHEET ==========

                    # Reload to access newly written sheet
                    book = load_workbook(backup_file_path)
                    sheet_result = book["Result"]

                    # Apply formatting...

                    for row in sheet_result.iter_rows(min_row=1, max_row=sheet_result.max_row,
                                                    min_col=1, max_col=sheet_result.max_column):
                        for cell in row:
                            cell.font = Font(name="Calibri", size=12)
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                            cell.border = Border(
                                left=Side(border_style=None),
                                right=Side(border_style=None),
                                top=Side(border_style=None),
                                bottom=Side(border_style=None)
                            )

                    # Format specific columns
                    # for col in ['A', 'N']:
                    #     for cell in sheet_result[col]:
                    #         # cell.number_format = '0'
                    #         cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)

                    for col in ['A', 'P']:
                        for cell in sheet_result[col]:
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

                    for col_idx in range(1, 16):
                        col_letter = get_column_letter(col_idx)
                        max_length = 0

                        # Loop through cells in the column to find the max length
                        for cell in sheet_result[col_letter]:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))

                        # Set column width
                        sheet_result.column_dimensions[col_letter].width = max_length + 2  # Adding a bit of padding

                    # Freeze first row
                    sheet_result.freeze_panes = sheet_result['A2']

                    # Save workbook
                    book.save(backup_file_path)

                    # Summary
                    end_time = datetime.datetime.now().time()
                    # Calculate the time difference in seconds
                    time_difference = datetime.datetime.combine(datetime.date.today(), end_time) - datetime.datetime.combine(datetime.date.today(), start_time)
                    # Extract the time difference in seconds
                    time_in_seconds = time_difference.total_seconds()

                    update_username(output_text, f"\n🎯 Script complete! {time_in_seconds}")
                    update_username(output_text, f"✅ Total Success: {success_count}")
                    update_username(output_text, f"❌ Total Failures: {failure_count}")
                    update_username(output_text, f"📄 Results saved to: {backup_file_path}")
                    messagebox.showinfo("Alert", f"Script complete")
                else:
                    missing = []
                    if "bahsegel1200 List" not in sheet_names:
                        missing.append("bahsegel1200 List")
                    messagebox.showinfo("Alert", f"❌ Missing Worksheet(s): {', '.join(missing)}")
            except Exception as e:
                messagebox.showinfo("Alert", f"❌ Error reading Excel file: {e}")
        else:
            messagebox.showinfo("Alert", f"bahsegel1200 File Not Found")
    threading.Thread(target=task).start()  

def rgb_to_hex(r, g, b):
    return f"#{r:02x}{g:02x}{b:02x}"

def write_output(text):
    output_text.config(state=tk.NORMAL)
    output_text.insert(tk.END, text)
    output_text.see(tk.END)
    output_text.config(state=tk.DISABLED)

def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    position_top = int(screen_height / 2 - height / 2)
    position_right = int(screen_width / 2 - width / 2)
    window.geometry(f'{width}x{height}+{position_right}+{position_top}')

root = tk.Tk()
root.title("Memon v.1.1")
root.geometry("500x400")
root.resizable(False, False)
center_window(root, 500, 400)
# Use raw string (r"...") or double backslashes
# root.iconbitmap("icon.ico")

button_font = ("Calibri", 12, "bold")
button_width = 15
button_height = 1

# Color Hex Codes from RGB
bahsegel1200_yellow = rgb_to_hex(255, 221, 0)      # #ffdd00
lowes_blue = rgb_to_hex(52, 130, 208)           # #0046ad
bahsegel1200_orange = rgb_to_hex(245, 130, 32)   # #f58220
stop_red = rgb_to_hex(240, 65, 48)   # #f58220

# Define border colors (as RGB hex)
bahsegel1200_border = rgb_to_hex(200, 170, 0)       # Golden
lowes_border = rgb_to_hex(0, 35, 100)          # Deep blue
bahsegel1200_border = rgb_to_hex(180, 80, 20)     # Deeper orange
stop_border = rgb_to_hex(255, 23, 0)     # Deeper orange

button_frame = tk.Frame(root)
button_frame.pack(pady=10)

# bahsegel1200
frame1 = tk.Frame(button_frame, bg=bahsegel1200_border, padx=2, pady=2)
frame1.pack(side=tk.LEFT, padx=5)

button1 = tk.Button(
    frame1, text="bahsegel1200", command=run_bahsegel1200_script,
    bg=bahsegel1200_yellow, fg="black", font=button_font,
    width=button_width, height=button_height,
    relief="flat", borderwidth=0
)
button1.pack()

# Create a Text widget to display username
output_text = tk.Text(root, height=20, wrap=tk.WORD, state=tk.DISABLED)
output_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

# Add a scrollbar to the Text widget
scrollbar = tk.Scrollbar(root, command=output_text.yview)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
output_text.config(yscrollcommand=scrollbar.set)

# Bind window close event to stop the task
root.protocol("WM_DELETE_WINDOW", on_closing)
root.mainloop()
