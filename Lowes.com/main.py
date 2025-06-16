import tkinter as tk
import threading
import json
import shutil
import datetime
import pandas as pd
import undetected_chromedriver as uc
import os
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from selenium.common.exceptions import NoSuchElementException, TimeoutException, StaleElementReferenceException, ElementNotInteractableException, ElementClickInterceptedException, NoSuchFrameException, NoSuchWindowException
from tkinter import messagebox

# Global flag to control whether the task is running
running = True
memon_kill = ""

def clean_sku(x):
    if pd.isna(x):
        return None
    try:
        if isinstance(x, float) and x.is_integer():
            return str(int(x))
        return str(x)
    except:
        return str(x)

def on_closing():
    global running
    running = False  # Stop the task
    root.quit()  # Close the Tkinter window

# Function to safely update the TextBox
def update_status(text_widget, message):
    # Use the after() method to safely update the widget in the main thread
    text_widget.after(0, lambda: text_widget.config(state=tk.NORMAL))  # Ensure widget is in NORMAL state
    text_widget.after(0, lambda: text_widget.insert(tk.END, message + "\n"))
    text_widget.after(0, lambda: text_widget.see(tk.END))
    text_widget.after(0, lambda: text_widget.config(state=tk.DISABLED))  # Disable the widget after update

def on_exit():
    global running
    running = False  # Stop the task
    root.quit()  # This ensures the program stops correctly

def run_lowes_script():
    global running
    
    # Clear the textbox when button is clicked
    output_text.config(state=tk.NORMAL)
    output_text.delete("1.0", tk.END)
    output_text.config(state=tk.DISABLED)
    
    def task():
        original_file_path = "Lowe's SKU.xlsx"
        # Check if the file exists
        if os.path.exists(original_file_path):
            try:
                workbook = load_workbook(original_file_path, read_only=True)
                sheet_names = workbook.sheetnames

                if "Lowe's SKU List" in sheet_names and "Discontinued SKU List" in sheet_names:
                    start_time = datetime.datetime.now().time()
                    update_status(output_text, f"Please wait while Chrome is starting up. (Lowe's) {start_time}")
                    # Setup WebDriver
                    # service = Service(ChromeDriverManager().install())

                    # Setup undetected Chrome
                    options = uc.ChromeOptions()
                    # Set up to block unnecessary background requests (external resources)
                    # options.add_argument("--headless")  # Run in headless mode (no GUI)
                    # options.add_argument("--headless=new")
                    # options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36")
                    options.add_argument("--disable-blink-features=AutomationControlled")  # Disable automation flag
                    options.add_argument("--disable-extensions")  # Disable extensions
                    options.add_argument("--disable-background-networking")  # Disable background networking
                    options.add_argument("--disable-software-rasterizer")  # Disable software rasterizer
                    options.add_argument("--disable-gpu")
                    # options.add_argument("--disable-http2")  # Force HTTP/1.1
                    options.add_argument("--no-sandbox")
                    options.add_argument("--log-level=3")
                    options.add_argument("--start-minimized")

                    # headers = {
                    #     "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                    #     "Accept": "application/json, text/plain, */*",
                    #     "Referer": "https://www.lowes.com/",
                    #     "Accept-Language": "en-US,en;q=0.9",
                    # }
                    
                    # driver = uc.Chrome(headers=headers, options=options)
                    driver = uc.Chrome(options=options)
                    update_status(output_text, f"Google Chrome has successfully launched.")

                    original_file_path = "Lowe's SKU.xlsx"

                    # Create timestamped backup
                    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                    backup_file_path = f"Lowe's SKU - Result {timestamp}.xlsx"
                    shutil.copy(original_file_path, backup_file_path)
                    # update_status(output_text, f"✅ Result created: {backup_file_path}")

                    # Load SKU list
                    df_sku = pd.read_excel(original_file_path, sheet_name="Lowe's SKU List")
                    df_discontinued = pd.read_excel(original_file_path, sheet_name="Discontinued SKU List")
                    # sku_list = df_sku.iloc[:, 0].dropna().apply(lambda x: str(int(x))).tolist()
                    search_list = df_sku.iloc[:, 0].apply(clean_sku).tolist()
                    sku_list = df_sku.iloc[:, 1].apply(clean_sku).tolist()
                    upc_list = df_sku.iloc[:, 3].apply(clean_sku).tolist()
                    discontinued_list = df_discontinued.iloc[:, 1].apply(clean_sku).tolist()

                    product_list = []
                    success_count = 0
                    failure_count = 0
                    remaining = len(sku_list)
                    memon = ""

                    for search, sku, upc in zip(search_list, sku_list, upc_list):
                            status = ""
                            buyable = ""
                            discontinued = "Yes" if upc in discontinued_list else "No"
                            retail_price = ""
                            out_of_stock = ""
                            quantity_available = 0
                            pd_url = ""
                            error_message = ""
                            access_text = ""
                            url = ""

                            if sku == None and search == None:
                                product_list.append([
                                    "SKU Not Found" if status == "" else "",
                                    "" if buyable == "" else "",
                                    discontinued,
                                    "" if retail_price == "" else "",
                                    "" if out_of_stock == "" else "",
                                    "" if quantity_available == 0 else "",
                                    "" if pd_url == ""else ""])
                                remaining -=1
                                update_status(output_text, f"❌ SKU {sku} Remaining {remaining}: Item/ProductID Not Found")
                                failure_count +=1
                            elif search == None:
                                product_list.append([
                                    "ProductID/Inactive Not Found" if status == "" else "",
                                    "" if buyable == "" else "",
                                    discontinued,
                                    "" if retail_price == "" else "",
                                    "" if out_of_stock == "" else "",
                                    "" if quantity_available == 0 else "",
                                    "" if pd_url == ""else ""])
                                remaining -=1
                                update_status(output_text, f"❌ SKU {sku} Remaining {remaining}: ProductID/Inactive Not Found")
                                failure_count +=1
                            else:
                                url = f"https://www.lowes.com/wpd/{search}/productdetail/undefined/Authenticated"
                                driver.get(url)

                                error_message = None
                                try:
                                    error_message_element = driver.find_element(By.XPATH, "//*[contains(text(), 'Looks Like This Page Is Missing or Moved')]")
                                    error_message = error_message_element.text
                                    # update_status(output_text, f"❌ Inactive")
                                except (NoSuchElementException, Exception) as e:
                                    # update_status(output_text, f"✅ Item Found")
                                    memon = ""

                                if access_text == "Access Denied":
                                    product_list.append([
                                        "Access Denied" if status == "" else "",
                                        "" if buyable == "" else "",
                                        discontinued,
                                        "" if retail_price == "" else "",
                                        "" if out_of_stock == "" else "",
                                        "" if quantity_available == 0 else "",
                                        "" if pd_url == ""else ""])
                                    remaining -=1
                                    failure_count +=1
                                    update_status(output_text, f"❌ SKU {sku} Remaining {remaining}: Access Denied")
                                elif error_message == "Looks Like This Page Is Missing or Moved":
                                # Regular expression to extract the ID between the last '/' and '?' in the URL
                                    product_list.append([
                                        "Inactive" if status == "" else "",
                                        "" if buyable == "" else "",
                                        discontinued,
                                        "" if retail_price == "" else "",
                                        "" if out_of_stock == "" else "",
                                        "" if quantity_available == 0 else "",
                                        "" if pd_url == ""else ""])
                                    remaining -=1
                                    failure_count +=1
                                    update_status(output_text, f"❌ SKU {sku} Remaining {remaining}: Inactive")
                                else:
                                    page_content = driver.page_source
                                    soup = BeautifulSoup(page_content, 'html.parser')
                                    pre_tag = soup.find("pre")

                                    if pre_tag:
                                        json_data = json.loads(pre_tag.text)
                                        product_details = json_data.get("productDetails", {})

                                        if not product_details:
                                            raise ValueError("No 'productDetails' found in JSON")

                                        product_key = list(product_details.keys())[0]
                                        product_data = product_details[product_key]
                                        product = product_data.get("product", {}) if isinstance(product_data.get("product", {}), dict) else {}

                                        item_inventory = product_data.get("itemInventory", {})
                                        item = item_inventory if isinstance(item_inventory, dict) else {}

                                        mfe_price = product_data.get("mfePrice", {})
                                        price = mfe_price if isinstance(mfe_price, dict) else {}

                                        analytics_data = item.get("analyticsData", {})
                                        parcel = analytics_data.get("parcel", {})
                                        additional_price = price.get("price", {})
                                        additional_data = additional_price.get("additionalData", {})

                                        raw_status = product.get("productStatus", "")
                                        status = "Active" if raw_status == True else "Inactive"
                                        # status = "Active"

                                        # raw_discontinued = product.get("productStatus", "")
                                        # discontinued = "Yes" if raw_discontinued == True else "No"

                                        raw_out_of_stock = item.get("totalAvailableQty", "")
                                        if raw_out_of_stock == "":
                                            out_of_stock = ""
                                        elif raw_out_of_stock == 0:
                                            out_of_stock = "Yes"
                                        else:
                                            out_of_stock = "No"

                                        # buyable = product.get("isBuyable", "")
                                        # if buyable == "Y":
                                        #     buyable = "Yes"
                                        # else:
                                        #     buyable = "No"

                                        if status == "Active" and out_of_stock == "No":
                                            buyable = "Yes"
                                        else:
                                            buyable = "No"

                                        retail_price = additional_data.get("retailPrice", "")
                                        quantity_available = item.get("totalAvailableQty", "")
                                        url = product.get("pdURL", "")
                                        pd_url = f"https://www.lowes.com{url}"

                                        product_list.append([
                                            status,
                                            buyable,
                                            discontinued,
                                            retail_price,
                                            out_of_stock,
                                            quantity_available,
                                            pd_url])
                                        remaining -=1
                                        success_count += 1
                                        update_status(output_text, f"✅ SKU {sku} Remaining {remaining}: Success")
                                    else:
                                        url = f"https://www.lowes.com/wpd/{search}/productdetail/undefined/Guest"
                                        driver.get(url)
                                        # time.sleep(1)

                                        page_content = driver.page_source
                                        soup = BeautifulSoup(page_content, 'html.parser')
                                        pre_tag = soup.find("pre")

                                        if pre_tag:
                                            json_data = json.loads(pre_tag.text)
                                            product_details = json_data.get("productDetails", {})

                                            if not product_details:
                                                raise ValueError("No 'productDetails' found in JSON")

                                            product_key = list(product_details.keys())[0]
                                            product_data = product_details[product_key]
                                            product = product_data.get("product", {}) if isinstance(product_data.get("product", {}), dict) else {}

                                            item_inventory = product_data.get("itemInventory", {})
                                            item = item_inventory if isinstance(item_inventory, dict) else {}

                                            mfe_price = product_data.get("mfePrice", {})
                                            price = mfe_price if isinstance(mfe_price, dict) else {}

                                            analytics_data = item.get("analyticsData", {})
                                            parcel = analytics_data.get("parcel", {})
                                            additional_price = price.get("price", {})
                                            additional_data = additional_price.get("additionalData", {})

                                            raw_status = product.get("productStatus", "")
                                            status = "Active" if raw_status == True else "Inactive"
                                            # status = "Active"

                                            # raw_discontinued = product.get("productStatus", "")
                                            # discontinued = "Yes" if raw_discontinued == True else "No"

                                            raw_out_of_stock = item.get("totalAvailableQty", "")
                                            if raw_out_of_stock == "":
                                                out_of_stock = ""
                                            elif raw_out_of_stock == 0:
                                                out_of_stock = "Yes"
                                            else:
                                                out_of_stock = "No"

                                            # buyable = product.get("isBuyable", "")
                                            # if buyable == "Y":
                                            #     buyable = "Yes"
                                            # else:
                                            #     buyable = "No"

                                            if status == "Active" and out_of_stock == "No":
                                                buyable = "Yes"
                                            else:
                                                buyable = "No"

                                            retail_price = additional_data.get("retailPrice", "")
                                            quantity_available = item.get("totalAvailableQty", "")
                                            url = product.get("pdURL", "")
                                            pd_url = f"https://www.lowes.com{url}"

                                            product_list.append([
                                                status,
                                                buyable,
                                                discontinued,
                                                retail_price,
                                                out_of_stock,
                                                quantity_available,
                                                pd_url])
                                            remaining -=1
                                            success_count += 1
                                            update_status(output_text, f"✅ SKU {sku} Remaining {remaining}: Success")
                                        else:
                                            product_list.append([
                                                "Access Denied" if status == "" else "",
                                                "" if buyable == "" else "",
                                                discontinued,
                                                "" if retail_price == "" else "",
                                                "" if out_of_stock == "" else "",
                                                "" if quantity_available == 0 else "",
                                                f"https://www.lowes.com/search?searchTerm={sku}" if pd_url == ""else ""])
                                            remaining -=1
                                            failure_count +=1
                                            update_status(output_text, f"❌ SKU {sku} Remaining {remaining}: Access Denied")
                            
                    # ✅ Quit browser after all SKUs are done
                    driver.quit()

                    # Load original sheet data (A to C)
                    book = load_workbook(backup_file_path)
                    df_existing = pd.read_excel(backup_file_path, sheet_name="Lowe's SKU List", usecols="A:D")

                    # Prepare result DataFrame
                    df_result = pd.DataFrame(product_list, columns=[
                        "Status", "Buyable", "Discontinued", "Retail Price", 
                        "Out Of Stock", "Quantity", "Product Link"
                    ])

                    # Combine and write to Excel
                    df_combined = pd.concat([df_existing, df_result], axis=1)

                    # ========== STEP 8: SAVE THE COMBINED DATA BACK TO THE BACKUP FILE ==========

                    if 'Result' in book.sheetnames:
                        del book['Result']
                        book.save(backup_file_path)  # Save after deleting to avoid error later

                    # Write new sheet
                    with pd.ExcelWriter(backup_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_combined.to_excel(writer, sheet_name="Result", index=False)

                    # ========== STEP 9: FORMAT THE SHEET ==========

                    # Reload to access newly written sheet
                    book = load_workbook(backup_file_path)
                    sheet_result = book["Result"]

                    # Apply formatting...

                    for row in sheet_result.iter_rows(min_row=1, max_row=sheet_result.max_row,
                                                    min_col=1, max_col=sheet_result.max_column):
                        for cell in row:
                            cell.font = Font(name="Calibri", size=12)
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                            cell.border = Border(
                                left=Side(border_style=None),
                                right=Side(border_style=None),
                                top=Side(border_style=None),
                                bottom=Side(border_style=None))

                    # Format specific columns
                    for col in ['A', 'D']:
                        for cell in sheet_result[col]:
                            cell.number_format = '0'
                            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)

                    for col in ['B', 'K']:
                        for cell in sheet_result[col]:
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

                    for col_idx in range(1, 11):
                        col_letter = get_column_letter(col_idx)
                        max_length = 0

                        # Loop through cells in the column to find the max length
                        for cell in sheet_result[col_letter]:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))

                        # Set column width
                        sheet_result.column_dimensions[col_letter].width = max_length + 2  # Adding a bit of padding

                    # Freeze first row
                    sheet_result.freeze_panes = sheet_result['A2']

                    # Save workbook
                    book.save(backup_file_path)

                    # Summary
                    end_time = datetime.datetime.now().time()
                    # Calculate the time difference in seconds
                    time_difference = datetime.datetime.combine(datetime.date.today(), end_time) - datetime.datetime.combine(datetime.date.today(), start_time)
                    # Extract the time difference in seconds
                    time_in_seconds = time_difference.total_seconds()
                    
                    update_status(output_text, f"\n🎯 Script complete! {time_in_seconds}")
                    update_status(output_text, f"✅ Total Success: {success_count}")
                    update_status(output_text, f"❌ Total Failures: {failure_count}")
                    update_status(output_text, f"📄 Results saved to: {backup_file_path}")
                    messagebox.showinfo("Alert", f"Script complete")
                else:
                    missing = []
                    if "Lowe's SKU List" not in sheet_names:
                        missing.append("Lowe's SKU List")
                    if "Discontinued SKU List" not in sheet_names:
                        missing.append("Discontinued SKU List")
                    messagebox.showinfo("Alert", f"❌ Missing Worksheet(s): {', '.join(missing)}")
            except Exception as e:
                messagebox.showinfo("Alert", f"❌ Error reading Excel file: {e}")
        else:
            messagebox.showinfo("Alert", f"Lowe's Sheet Not Found")

    threading.Thread(target=task).start()    

def run_lowesID_script():
    global running
    
    # Clear the textbox when button is clicked
    output_text.config(state=tk.NORMAL)
    output_text.delete("1.0", tk.END)
    output_text.config(state=tk.DISABLED)
    
    def task():
        original_file_path = "Lowe's SKU.xlsx"
        # Check if the file exists
        if os.path.exists(original_file_path):
            try:
                workbook = load_workbook(original_file_path, read_only=True)
                sheet_names = workbook.sheetnames

                if "Lowe's SKU List" in sheet_names:
                    start_time = datetime.datetime.now().time()
                    update_status(output_text, f"Please wait while Chrome is starting up. (Lowe's) {start_time}")
                    # Setup WebDriver
                    # service = Service(ChromeDriverManager().install())

                    # Setup undetected Chrome
                    options = uc.ChromeOptions()
                    # Set up to block unnecessary background requests (external resources)
                    # options.add_argument("--headless")  # Run in headless mode (no GUI)
                    # options.add_argument("--headless=new")
                    # ooptions.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36")
                    options.add_argument("--disable-blink-features=AutomationControlled")  # Disable automation flag
                    options.add_argument("--disable-extensions")  # Disable extensions
                    options.add_argument("--disable-background-networking")  # Disable background networking
                    options.add_argument("--disable-software-rasterizer")  # Disable software rasterizer
                    options.add_argument("--disable-gpu")
                    # options.add_argument("--disable-http2")  # Force HTTP/1.1
                    options.add_argument("--no-sandbox")
                    options.add_argument("--log-level=3")
                    options.add_argument("--start-minimized")

                    # headers = {
                    #     "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                    #     "Accept": "application/json, text/plain, */*",
                    #     "Referer": "https://www.lowes.com/",
                    #     "Accept-Language": "en-US,en;q=0.9",
                    # }
                    
                    # driver = uc.Chrome(headers=headers, options=options)
                    driver = uc.Chrome(options=options)
                    update_status(output_text, f"Google Chrome has successfully launched.")

                    original_file_path = "Lowe's SKU.xlsx"

                    # Create timestamped backup
                    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                    backup_file_path = f"Lowe's SKU - Result {timestamp}.xlsx"
                    shutil.copy(original_file_path, backup_file_path)
                    # update_status(output_text, f"✅ Result created: {backup_file_path}")

                    # Load SKU list
                    df_sku = pd.read_excel(original_file_path, sheet_name="Lowe's SKU List")
                    # sku_list = df_sku.iloc[:, 0].dropna().apply(lambda x: str(int(x))).tolist()
                    sku_list = df_sku.iloc[:, 3].apply(clean_sku).tolist()
                    upc_list = df_sku.iloc[:, 0].apply(clean_sku).tolist()

                    product_list = []
                    success_count = 0
                    failure_count = 0
                    memon = ""
                    remaining = len(sku_list)

                    for sku, upc in zip(sku_list, upc_list):
                            status = ""
                            error_message = ""
                            sku_id = ""
                            access_text = ""
                            url = ""

                            if upc == None:
                                url = f"https://www.lowes.com/search?searchTerm={sku}"
                                driver.get(url)

                                try:
                                    access_element = driver.find_element(By.TAG_NAME, "h1")
                                    if access_element.text.strip() == "Access Denied":
                                        access_text = "Access Denied"
                                    else:
                                        # Define a custom condition
                                        def url_or_element(driver):
                                            if driver.current_url != url:
                                                return True
                                            try:
                                                # Try to locate the element
                                                driver.find_element(By.CSS_SELECTOR, 'h1.styles__H1-sc-11vpuyu-0.krJSUv.typography.variant--h1.align--left.heading.bold')
                                                return True
                                            except:
                                                return False
                                        # Wait for either the URL to change or the element to appear
                                        WebDriverWait(driver, 10).until(url_or_element)

                                        # WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'h1.styles__H1-sc-11vpuyu-0.krJSUv.typography.variant--h1.align--left.heading.bold')))
                                        error_message = None
                                        try:
                                            error_message_element = driver.find_element(By.XPATH, "//*[contains(text(), 'Try a different search or one of these search terms instead:')]")
                                            error_message = error_message_element.text
                                            # update_status(output_text, f"❌ Inactive")
                                        except:
                                            # update_status(output_text, f"✅ Item Found")
                                            memon = ""
                                            WebDriverWait(driver, 10).until(lambda driver: driver.current_url != url)

                                        sku_id = driver.current_url.split('/')[-1].split('?')[0]
                                except:
                                # Define a custom condition
                                    def url_or_element(driver):
                                        if driver.current_url != url:
                                            return True
                                        try:
                                            # Try to locate the element
                                            driver.find_element(By.CSS_SELECTOR, 'h1.styles__H1-sc-11vpuyu-0.krJSUv.typography.variant--h1.align--left.heading.bold')
                                            return True
                                        except:
                                            return False
                                    # Wait for either the URL to change or the element to appear
                                    WebDriverWait(driver, 10).until(url_or_element)

                                    # WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'h1.styles__H1-sc-11vpuyu-0.krJSUv.typography.variant--h1.align--left.heading.bold')))
                                    error_message = None
                                    try:
                                        error_message_element = driver.find_element(By.XPATH, "//*[contains(text(), 'Try a different search or one of these search terms instead:')]")
                                        error_message = error_message_element.text
                                        # update_status(output_text, f"❌ Inactive")
                                    except Exception as e:
                                        # update_status(output_text, f"✅ Item Found")
                                        memon = ""
                                    WebDriverWait(driver, 10).until(lambda driver: driver.current_url != url)

                                    sku_id = driver.current_url.split('/')[-1].split('?')[0]

                                if access_text == "Access Denied":
                                    product_list.append([
                                        "Access Denied" if status == "" else "",])
                                    remaining -=1
                                    failure_count +=1
                                    update_status(output_text, f"❌ SKU {sku} Remaining {remaining}: Access Denied")
                                elif error_message == "Try a different search or one of these search terms instead:":
                                # Regular expression to extract the ID between the last '/' and '?' in the URL
                                    product_list.append([
                                        "Not Found" if status == "" else "",])
                                    remaining -=1
                                    failure_count +=1
                                    update_status(output_text, f"❌ SKU {sku} Remaining {remaining}: Not Found")
                                elif sku_id == "search":
                                    search_term = url.split('=')[-1]  # Get the search term from the URL

                                    WebDriverWait(driver, 10).until(
                                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, '[data-itemnumber]')))

                                    elements = driver.find_elements(By.CSS_SELECTOR, '[data-itemnumber]')
                                    found = False
                                    for index, element in enumerate(elements):
                                        data_itemnumber = element.get_attribute('data-itemnumber')
                                        
                                        # Check if the data-itemnumber matches the search term
                                        if data_itemnumber == search_term:
                                            data_id = element.get_attribute('data-id')
                                            # update_status(output_text, f"Found matching element: data-itemnumber = {data_itemnumber}, data-id = {data_id}")
                                            found = True
                                            break
                                        else:
                                            update_status(output_text, f"Element {index}: data-itemnumber = {data_itemnumber} doesn't match search term.")
                                    if not found:
                                        update_status(output_text, f"No matching elements found for search term {search_term}.")

                                    product_list.append([
                                        data_id])
                                    remaining -=1
                                    success_count += 1
                                    update_status(output_text, f"✅ SKU {sku} Remaining {remaining}: Success")
                                else:
                                    product_list.append([
                                        sku_id])
                                    remaining -=1
                                    success_count += 1
                                    update_status(output_text, f"✅ SKU {sku} Remaining {remaining}: Success")
                            else:
                                product_list.append([
                                    "Already Product Found" if status == "" else "",])
                                remaining -=1
                                update_status(output_text, f"❌ SKU {sku} Remaining {remaining}: Already Product Found")
                                failure_count +=1


                    # ✅ Quit browser after all SKUs are done
                    driver.quit()

                    # Load original sheet data (A to C)
                    book = load_workbook(backup_file_path)
                    df_existing = pd.read_excel(backup_file_path, sheet_name="Lowe's SKU List", usecols="A:D")

                    # Prepare result DataFrame
                    df_result = pd.DataFrame(product_list, columns=[
                        "Get Product ID"
                    ])

                    # Combine and write to Excel
                    df_combined = pd.concat([df_existing, df_result], axis=1)

                    # ========== STEP 8: SAVE THE COMBINED DATA BACK TO THE BACKUP FILE ==========

                    if 'Result' in book.sheetnames:
                        del book['Result']
                        book.save(backup_file_path)  # Save after deleting to avoid error later

                    # Write new sheet
                    with pd.ExcelWriter(backup_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_combined.to_excel(writer, sheet_name="Result", index=False)

                    # ========== STEP 9: FORMAT THE SHEET ==========

                    # Reload to access newly written sheet
                    book = load_workbook(backup_file_path)
                    sheet_result = book["Result"]

                    # Apply formatting...

                    for row in sheet_result.iter_rows(min_row=1, max_row=sheet_result.max_row,
                                                    min_col=1, max_col=sheet_result.max_column):
                        for cell in row:
                            cell.font = Font(name="Calibri", size=12)
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                            cell.border = Border(
                                left=Side(border_style=None),
                                right=Side(border_style=None),
                                top=Side(border_style=None),
                                bottom=Side(border_style=None))

                    # Format specific columns
                    for col in ['A', 'E']:
                        for cell in sheet_result[col]:
                            cell.number_format = '0'
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

                    for col_idx in range(1, 6):
                        col_letter = get_column_letter(col_idx)
                        max_length = 0

                        # Loop through cells in the column to find the max length
                        for cell in sheet_result[col_letter]:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))

                        # Set column width
                        sheet_result.column_dimensions[col_letter].width = max_length + 2  # Adding a bit of padding

                    # Freeze first row
                    sheet_result.freeze_panes = sheet_result['A2']

                    # Save workbook
                    book.save(backup_file_path)

                    # Summary
                    end_time = datetime.datetime.now().time()
                    # Calculate the time difference in seconds
                    time_difference = datetime.datetime.combine(datetime.date.today(), end_time) - datetime.datetime.combine(datetime.date.today(), start_time)
                    # Extract the time difference in seconds
                    time_in_seconds = time_difference.total_seconds()
                    
                    update_status(output_text, f"\n🎯 Script complete! {time_in_seconds}")
                    update_status(output_text, f"✅ Total Success: {success_count}")
                    update_status(output_text, f"❌ Total Failures: {failure_count}")
                    update_status(output_text, f"📄 Results saved to: {backup_file_path}")
                    messagebox.showinfo("Alert", f"Script complete")
                else:
                    missing = []
                    if "Lowe's SKU List" not in sheet_names:
                        missing.append("Lowe's SKU List")
                    messagebox.showinfo("Alert", f"❌ Missing Worksheet(s): {', '.join(missing)}")
            except Exception as e:
                messagebox.showinfo("Alert", f"❌ Error reading Excel file: {e}")
        else:
            messagebox.showinfo("Alert", f"Lowe's Sheet Not Found")

    threading.Thread(target=task).start()

def rgb_to_hex(r, g, b):
    return f"#{r:02x}{g:02x}{b:02x}"

def write_output(text):
    output_text.config(state=tk.NORMAL)
    output_text.insert(tk.END, text)
    output_text.see(tk.END)
    output_text.config(state=tk.DISABLED)

def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    position_top = int(screen_height / 2 - height / 2)
    position_right = int(screen_width / 2 - width / 2)
    window.geometry(f'{width}x{height}+{position_right}+{position_top}')

root = tk.Tk()
root.title("Memon Lowe's v.1.15")
root.geometry("500x400")
root.resizable(False, False)
center_window(root, 500, 400)
# Use raw string (r"...") or double backslashes
# icon_path = os.path.join(os.path.dirname(__file__), "logo.ico")
# root.iconbitmap(icon_path)

button_font = ("Calibri", 12, "bold")
button_width = 15
button_height = 1

# Color Hex Codes from RGB
lowes_blue = rgb_to_hex(52, 130, 208)           # #0046ad
stop_red = rgb_to_hex(240, 65, 48)   # #f58220

# Define border colors (as RGB hex)
lowes_border = rgb_to_hex(0, 35, 100)          # Deep blue
stop_border = rgb_to_hex(255, 23, 0)     # Deeper orange

button_frame = tk.Frame(root)
button_frame.pack(pady=10)

# Lowes
frame2 = tk.Frame(button_frame, bg=lowes_border, padx=2, pady=2)
frame2.pack(side=tk.LEFT, padx=5)

button2 = tk.Button(
    frame2, text="Lowe's SKU", command=run_lowes_script,
    bg=lowes_blue, fg="white", font=button_font,
    width=button_width, height=button_height,
    relief="flat", borderwidth=0
)
button2.pack()

# Stop
frame4 = tk.Frame(button_frame, bg=stop_border, padx=2, pady=2)
frame4.pack(side=tk.LEFT, padx=5)

button4 = tk.Button(
    frame4, text="ID", command=run_lowesID_script,
    bg=stop_red, fg="white", font=button_font,
    width=2, height=button_height,
    relief="flat", borderwidth=0
)
button4.pack()

# Create a Text widget to display status
output_text = tk.Text(root, height=20, wrap=tk.WORD, state=tk.DISABLED)
output_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

# Add a scrollbar to the Text widget
scrollbar = tk.Scrollbar(root, command=output_text.yview)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
output_text.config(yscrollcommand=scrollbar.set)

# Bind window close event to stop the task
root.protocol("WM_DELETE_WINDOW", on_closing)
root.mainloop()