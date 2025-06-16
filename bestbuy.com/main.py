import tkinter as tk
import threading
import json
import shutil
import datetime
import pandas as pd
import undetected_chromedriver as uc
import requests
import time
import concurrent.futures
import os
import uuid
import random
import string
import queue
import traceback
import queue
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from selenium.common.exceptions import NoSuchElementException, TimeoutException, StaleElementReferenceException, ElementNotInteractableException, ElementClickInterceptedException, NoSuchFrameException, NoSuchWindowException
from tkinter import messagebox

# Global flag to control whether the task is running
running = True

def clean_sku(x):
    if pd.isna(x):
        return None
    try:
        if isinstance(x, float) and x.is_integer():
            return str(int(x))
        return str(x)
    except:
        return str(x)

def on_closing():
    global running
    running = False  # Stop the task
    root.quit()  # Close the Tkinter window

# Function to safely update the TextBox
def update_status(text_widget, message):
    # Use the after() method to safely update the widget in the main thread
    text_widget.after(0, lambda: text_widget.config(state=tk.NORMAL))  # Ensure widget is in NORMAL state
    text_widget.after(0, lambda: text_widget.insert(tk.END, message + "\n"))
    text_widget.after(0, lambda: text_widget.see(tk.END))
    text_widget.after(0, lambda: text_widget.config(state=tk.DISABLED))  # Disable the widget after update

def on_exit():
    global running
    running = False  # Stop the task
    root.quit()  # This ensures the program stops correctly

def run_bestbuy_script():
    global running
    # Clear the textbox when button is clicked
    output_text.config(state=tk.NORMAL)
    output_text.delete("1.0", tk.END)
    output_text.config(state=tk.DISABLED)

    def task():
        original_file_path = "BestBuy SKU.xlsx"
        # Check if the file exists
        if os.path.exists(original_file_path):
            try:
                workbook = load_workbook(original_file_path, read_only=True)
                sheet_names = workbook.sheetnames

                if "Best Buy SKU List" in sheet_names and "Discontinued SKU List" in sheet_names:
                    code()
                else:
                    missing = []
                    if "Best Buy SKU List" not in sheet_names:
                        missing.append("Best Buy SKU List")
                    if "Discontinued SKU List" not in sheet_names:
                        missing.append("Discontinued SKU List")
                    messagebox.showinfo("Alert", f"❌ Missing Worksheet(s): {', '.join(missing)}")
            except Exception as e:
                workbook = load_workbook(original_file_path, read_only=True)
                sheet_names = workbook.sheetnames

                if "Best Buy SKU List" in sheet_names and "Discontinued SKU List" in sheet_names:
                    code()
                else:
                    missing = []
                    if "Best Buy SKU List" not in sheet_names:
                        missing.append("Best Buy SKU List")
                    if "Discontinued SKU List" not in sheet_names:
                        missing.append("Discontinued SKU List")
                    messagebox.showinfo("Alert", f"❌ Missing Worksheet(s): {', '.join(missing)}")
        else:
            messagebox.showinfo("Alert", f"BestBuy Sheet Not Found")

    def code():
            # The endpoint URL for BestBuy's GraphQL API
            url = "https://www.bestbuy.com/gateway/graphql"

            # Headers to mimic a real browser request
            headers = {
                "Accept": "application/json",
                "Content-Type": "application/json",
                "Origin": "https://www.bestbuy.com",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
            }

            # GraphQL query to fetch product data
            # Updated GraphQL query to use `country` instead of `region`
            query = """
            query MyQuery($skuId: String!, $storeId: String!) {
            productBySkuId(skuId: $skuId) {
                buyingOptions {
                skuId
                pdpUrl
                code
                type
                product {
                    price(input: {salesChannel: "Web"}) {
                    customerPrice
                    }
                    fulfillmentOptions(input: {inStorePickup: {storeId: $storeId, searchNearby: true}}) {
                    ispuDetails {
                        maximumDistanceChecked
                        numberOfStoresChecked
                        ispuAvailability {
                        instoreInventoryAvailable
                        quantity
                        minPickupInHours
                        maxDate
                        }
                        nearbyLocation {
                        availability {
                            minPickupInHours
                            quantity
                            maxDate
                            instoreInventoryAvailable
                        }
                        }
                    }
                    shippingDetails {
                        destinationZipCode
                        shippingAvailability {
                        shippingEligible
                        customerLOSGroup {
                            customerLosGroupId
                        }
                        }
                    }
                    }
                }
                }
            }
            }
            """
            # def collect_all_quantities_with_zip(data):
            #     quantities_info = []

            #     # Build locationId → zipCode map
            #     location_zipcode_map = {}
            #     for loc in data.get("ispu", {}).get("locations", []):
            #         if 'id' in loc and 'zipCode' in loc:
            #             location_zipcode_map[str(loc["id"])] = loc["zipCode"]

            #     # Get the first item from ispu > items
            #     items = data.get("ispu", {}).get("items", [])
            #     if not items:
            #         return quantities_info  # no items, return empty list

            #     # Loop through locations under the first item
            #     for loc in items[0].get("locations", []):
            #         location_id = str(loc.get("locationId", ""))
            #         if not location_id:
            #             continue

            #         availability = loc.get("availability", {})
            #         zip_code = location_zipcode_map.get(location_id)

            #         for key, value in availability.items():
            #             if "quantity" in key.lower() and isinstance(value, (int, float)):
            #                 quantities_info.append({
            #                     "locationId": location_id,
            #                     "zipCode": zip_code,
            #                     "quantity": value
            #                 })

            #     return quantities_info
            
            # def collect_all_graphquantities(data):
            #     graphquantities = []

            #     def recursive_search(d):
            #         if isinstance(d, dict):
            #             for key, value in d.items():
            #                 if 'quantity' in key.lower():
            #                     # Convert value to int safely
            #                     try:
            #                         num = int(value) if value is not None else 0
            #                         graphquantities.append(num)
            #                     except (ValueError, TypeError):
            #                         graphquantities.append(0)  # fallback if value is not a valid number
            #                 else:
            #                     recursive_search(value)
            #         elif isinstance(d, list):
            #             for item in d:
            #                 recursive_search(item)

            #     recursive_search(data)
            #     return graphquantities

            start_time = datetime.datetime.now().time()
            update_status(output_text, f"Please wait while API running. {start_time}")

            # ========== STEP 1: BACKUP ORIGINAL FILE ==========
            original_file_path = "BestBuy SKU.xlsx"
            # originalstore_file_path = "BestBuy Storelist v.1.xlsx"

            # Create timestamped backup
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
            backup_file_path = f"BestBuy SKU - Result {timestamp}.xlsx"
            shutil.copy(original_file_path, backup_file_path)
            # update_status(output_text, f"✅ Result created: {backup_file_path}")

            # ========== STEP 2: LOAD SKUs FROM ORIGINAL FILE ==========
            df_sku = pd.read_excel(original_file_path, sheet_name="Best Buy SKU List")
            df_discontinued = pd.read_excel(original_file_path, sheet_name="Discontinued SKU List")
            # df_store = pd.read_excel(originalstore_file_path, sheet_name="Best Buy Store List")
            
            # sku_list = df_sku.iloc[:, 0].dropna().apply(lambda x: str(int(x))).tolist()
            sku_list = df_sku.iloc[:, 0].apply(clean_sku).tolist()
            upc_list = df_sku.iloc[:, 2].apply(clean_sku).tolist()
            discontinued_list = df_discontinued.iloc[:, 1].apply(clean_sku).tolist()
            # storeId_list = df_store.iloc[:, 0].apply(clean_sku).tolist()
            # zipCode_list = df_store.iloc[:, 5].apply(clean_sku).tolist()

            # ========== STEP 3: SETUP ==========

            api_key = "lwp7anABJejNPJR3WAFLITOK"
            product_list = []
            success_count = 0
            failure_count = 0
            remaining = len(sku_list)
            # zip_code = []
            
            # ========== STEP 4: PROCESS EACH SKU ==========

            for sku, upc in zip(sku_list, upc_list):
                status = ""
                buyable = ""
                discontinued = "Yes" if upc in discontinued_list else "No"
                retail_price = ""
                out_of_stock = ""
                quantity_available = 0
                pd_url = ""
                # remainingzip = len(zipCode_list)

                if sku == None:
                    product_list.append([
                        "SKU Not Found" if status == "" else "",
                        "" if buyable == "" else "",
                        discontinued,
                        "" if retail_price == "" else "",
                        "" if out_of_stock == "" else "",
                        "" if quantity_available == 0 else "",
                        "" if pd_url == ""else ""])
                    remaining -=1
                    update_status(output_text, f"❌ SKU {sku} Remaining {remaining}: Not Found")
                    failure_count +=1
                else:
                    api_url = (
                        f"https://api.bestbuy.com/v1/products/{sku}.json"
                        f"?show=active,orderable,regularPrice,onlineAvailability,quantityLimit,url"
                        f"&apiKey={api_key}"
                    )

                    time.sleep(1)  # Respect API limits
                    # update_status(output_text, f"🔍 Processing SKU: {sku}...")

                    session = requests.session()
                    response = requests.get(api_url, timeout=10)
                    if response.status_code == 200:
                        data = response.json()

                        # Extract status details
                        raw_status = data.get("active", "")
                        status = "Active" if raw_status == True else "Inactive"

                        # raw_discontinued = data.get("onlineAvailability", "")
                        # discontinued = "Yes" if raw_discontinued == True else "No"

                        # if status == "Active" and discontinued == "No":
                        #     buyable = "Yes"
                        # else:
                        #     buyable = "No"

                        raw_buyable = data.get("orderable", "")
                        buyable = "No" if raw_buyable == "NotOrderable" else "Yes"

                        retail_price = data.get("regularPrice", "")
                        pd_url = data.get("url") or f"https://api.bestbuy.com/click/-/{sku}/pdp"

                        # # API URL for BestBuy's Store Availability API
                        # url = 'https://www.bestbuy.com/productfulfillment/c/api/2.0/storeAvailability'

                        # # Headers to mimic a real browser request
                        # headers = {
                        #     "Authority": "www.bestbuy.com",
                        #     "Method": "POST",
                        #     "Scheme": "https",
                        #     "Accept": "application/json",
                        #     "Accept-Encoding": "gzip, deflate, br, zstd",
                        #     "Accept-Language": "en-US,en;q=0.9",
                        #     "Content-Type": "application/json",
                        #     "Origin": "https://www.bestbuy.com",
                        #     "Priority": "u=1, i",
                        #     "Connection": "keep-alive",
                        #     "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
                        # }

                        # # Main logic for handling zip codes and collecting quantities
                        # quantity_available = 0

                        # def is_valid_zip(store_zip, zip_filter):
                        #     # If zip_code is blank, consider it a valid match (should process)
                        #     if not zip_filter:
                        #         return True  # Process if no zip_code is provided
                            
                        #     # If input is string (e.g. from text input), split into list
                        #     if isinstance(zip_filter, str):
                        #         zip_codes = zip_filter.split(',')
                        #     else:
                        #         zip_codes = list(zip_filter)  # Accepts set, list, etc.
                            
                        #     # If store_zip is in zip_codes, return False (i.e., do not process)
                        #     return store_zip not in zip_codes  # Invert logic here

                        # # Initialize sets
                        # unique_valid_zips = set()         # ZIPs that passed is_valid_zip
                        # collected_zip_codes = set()       # ZIPs collected from the API response

                        # for store_zip in set(zipCode_list):
                        #     # Check if zip_code filter is set and valid for this store_zip
                        #     should_process = False
                        #     if collected_zip_codes:  # zip_code is not blank
                        #         if is_valid_zip(store_zip, collected_zip_codes):
                        #             should_process = True  # Only process if store_zip does NOT match zip_code
                        #             remainingzip -= 1
                        #     else:  # zip_code is blank
                        #         should_process = True

                        #     if should_process:
                        #         unique_valid_zips.add(store_zip)
                        #         data = {
                        #             "zipCode": store_zip,
                        #             # "showOnShelf": False,
                        #             # "lookupInStoreQuantity": True,
                        #             # "xboxAllAccess": True,
                        #             # "consolidated": False,
                        #             # "showOnlyOnShelf": False,
                        #             # "showInStore": False,
                        #             # "pickupTypes": ["CURBSIDE"],
                        #             # "onlyBestBuyLocations": False,
                        #             "items": [
                        #                 {
                        #                     "sku": sku,
                        #                     "condition": None,
                        #                     "reservationToken": None,
                        #                     "selectedServices": [],
                        #                     "requiredAccessories": [],
                        #                     "isTradeIn": False,
                        #                     "isLeased": False
                        #                 }
                        #             ]
                        #         }

                        #         try:
                        #             response = requests.post(url, json=data, headers=headers, timeout=10)
                        #         except:
                        #             response = requests.post(url, json=data, headers=headers, timeout=10)

                        #         if response.status_code == 200:
                        #             data = response.json()
                        #             quantities_info = collect_all_quantities_with_zip(data)

                        #             location_map = {}
                        #             for item in quantities_info:
                        #                 location_id = item['locationId']
                        #                 available_quantity = item['quantity']
                        #                 zip_code = item.get('zipCode') or ""

                        #                 collected_zip_codes.add(zip_code)

                        #                 if location_id in location_map:
                        #                     if location_map[location_id]['quantity'] == 0 and available_quantity > 0:
                        #                         location_map[location_id]['quantity'] = available_quantity
                        #                 else:
                        #                     if available_quantity > 0:
                        #                         location_map[location_id] = {
                        #                             'quantity': available_quantity,
                        #                             'zipCode': zip_code
                        #                         }

                        #             # Calculate total available quantity
                        #             total_quantity = sum(loc['quantity'] for loc in location_map.values())
                        #             quantity_available += total_quantity
                        #             remainingzip -= 1
                        #             update_status(output_text, f"✅ SKU {sku} Remaining {remaining}: Zipcode {store_zip} Remaining {remainingzip}: Response Success")
                        #         else:
                        #             remainingzip -= 1
                        #             update_status(output_text, f"❌ SKU {sku} Remaining {remaining}: Zipcode {store_zip} Remaining {remainingzip}: Response Failed")

                        # if quantity_available == 0:
                        #     quantity_available = "N/A"
                        # elif quantity_available == "":
                        #     quantity_available = "N/A"
                        # else: 
                        #     quantity_available
                        quantity_available = ""

                        # if quantity_available == "N/A":
                        #     out_of_stock = ""
                        # elif quantity_available == 0:
                        #     out_of_stock = "Yes"
                        # else:
                        #     out_of_stock = "No"
                        out_of_stock = ""

                        product_list.append([ 
                            status,
                            buyable,
                            discontinued,
                            retail_price,
                            out_of_stock,                
                            quantity_available,
                            pd_url])
                        remaining -=1
                        success_count += 1
                        update_status(output_text, f"✅ SKU {sku} Remaining {remaining}: Success")
                    else:
                        api_url = (
                            f"https://api.bestbuy.com/v1/products/{sku}.json"
                            f"?show=active,orderable,regularPrice,onlineAvailability,quantityLimit,url"
                            f"&apiKey={api_key}"
                        )

                        time.sleep(1)  # Respect API limits
                        # update_status(output_text, f"🔍 Processing SKU: {sku}...")

                        session = requests.session()
                        response = requests.get(api_url, timeout=10)
                        if response.status_code == 200:
                            data = response.json()

                            # Extract status details
                            raw_status = data.get("active", "")
                            status = "Active" if raw_status == True else "Inactive"

                            # raw_discontinued = data.get("onlineAvailability", "")
                            # discontinued = "Yes" if raw_discontinued == True else "No"

                            # if status == "Active" and discontinued == "No":
                            #     buyable = "Yes"
                            # else:
                            #     buyable = "No"

                            raw_buyable = data.get("orderable", "")
                            buyable = "No" if raw_buyable == "NotOrderable" else "Yes"

                            retail_price = data.get("regularPrice", "")
                            pd_url = data.get("url") or f"https://api.bestbuy.com/click/-/{sku}/pdp"

                            # # API URL for BestBuy's Store Availability API
                            # url = 'https://www.bestbuy.com/productfulfillment/c/api/2.0/storeAvailability'

                            # # Headers to mimic a real browser request
                            # headers = {
                            #     "Authority": "www.bestbuy.com",
                            #     "Method": "POST",
                            #     "Scheme": "https",
                            #     "Accept": "application/json",
                            #     "Accept-Encoding": "gzip, deflate, br, zstd",
                            #     "Accept-Language": "en-US,en;q=0.9",
                            #     "Content-Type": "application/json",
                            #     "Origin": "https://www.bestbuy.com",
                            #     "Priority": "u=1, i",
                            #     "Connection": "keep-alive",
                            #     "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
                            # }

                            # # Main logic for handling zip codes and collecting quantities
                            # quantity_available = 0

                            # def is_valid_zip(store_zip, zip_filter):
                            #     # If zip_code is blank, consider it a valid match (should process)
                            #     if not zip_filter:
                            #         return True  # Process if no zip_code is provided
                                
                            #     # If input is string (e.g. from text input), split into list
                            #     if isinstance(zip_filter, str):
                            #         zip_codes = zip_filter.split(',')
                            #     else:
                            #         zip_codes = list(zip_filter)  # Accepts set, list, etc.
                                
                            #     # If store_zip is in zip_codes, return False (i.e., do not process)
                            #     return store_zip not in zip_codes  # Invert logic here

                            # # Initialize sets
                            # unique_valid_zips = set()         # ZIPs that passed is_valid_zip
                            # collected_zip_codes = set()       # ZIPs collected from the API response

                            # for store_zip in set(zipCode_list):
                            #     # Check if zip_code filter is set and valid for this store_zip
                            #     should_process = False
                            #     if collected_zip_codes:  # zip_code is not blank
                            #         if is_valid_zip(store_zip, collected_zip_codes):
                            #             should_process = True  # Only process if store_zip does NOT match zip_code
                            #             remainingzip -= 1
                            #     else:  # zip_code is blank
                            #         should_process = True

                            #     if should_process:
                            #         unique_valid_zips.add(store_zip)
                            #         data = {
                            #             "zipCode": store_zip,
                            #             # "showOnShelf": False,
                            #             # "lookupInStoreQuantity": True,
                            #             # "xboxAllAccess": True,
                            #             # "consolidated": False,
                            #             # "showOnlyOnShelf": False,
                            #             # "showInStore": False,
                            #             # "pickupTypes": ["CURBSIDE"],
                            #             # "onlyBestBuyLocations": False,
                            #             "items": [
                            #                 {
                            #                     "sku": sku,
                            #                     "condition": None,
                            #                     "reservationToken": None,
                            #                     "selectedServices": [],
                            #                     "requiredAccessories": [],
                            #                     "isTradeIn": False,
                            #                     "isLeased": False
                            #                 }
                            #             ]
                            #         }

                            #         try:
                            #             response = requests.post(url, json=data, headers=headers, timeout=10)
                            #         except:
                            #             response = requests.post(url, json=data, headers=headers, timeout=10)

                            #         if response.status_code == 200:
                            #             data = response.json()
                            #             quantities_info = collect_all_quantities_with_zip(data)

                            #             location_map = {}
                            #             for item in quantities_info:
                            #                 location_id = item['locationId']
                            #                 available_quantity = item['quantity']
                            #                 zip_code = item.get('zipCode') or ""

                            #                 collected_zip_codes.add(zip_code)

                            #                 if location_id in location_map:
                            #                     if location_map[location_id]['quantity'] == 0 and available_quantity > 0:
                            #                         location_map[location_id]['quantity'] = available_quantity
                            #                 else:
                            #                     if available_quantity > 0:
                            #                         location_map[location_id] = {
                            #                             'quantity': available_quantity,
                            #                             'zipCode': zip_code
                            #                         }

                            #             # Calculate total available quantity
                            #             total_quantity = sum(loc['quantity'] for loc in location_map.values())
                            #             quantity_available += total_quantity
                            #             remainingzip -= 1
                            #             update_status(output_text, f"✅ SKU {sku} Remaining {remaining}: Zipcode {store_zip} Remaining {remainingzip}: Response Success")
                            #         else:
                            #             remainingzip -= 1
                            #             update_status(output_text, f"❌ SKU {sku} Remaining {remaining}: Zipcode {store_zip} Remaining {remainingzip}: Response Failed")

                            # if quantity_available == 0:
                            #     quantity_available = "N/A"
                            # elif quantity_available == "":
                            #     quantity_available = "N/A"
                            # else: 
                            #     quantity_available
                            quantity_available = ""

                            # if quantity_available == "N/A":
                            #     out_of_stock = ""
                            # elif quantity_available == 0:
                            #     out_of_stock = "Yes"
                            # else:
                            #     out_of_stock = "No"
                            out_of_stock = ""

                            product_list.append([ 
                                status,
                                buyable,
                                discontinued,
                                retail_price,
                                out_of_stock,                
                                quantity_available,
                                pd_url])
                            remaining -=1
                            success_count += 1
                            update_status(output_text, f"✅ SKU {sku} Remaining {remaining}: Success")
                        else:
                            product_list.append([
                                "Inactive" if status == "" else "",
                                "" if buyable == "" else "",
                                discontinued,
                                "" if retail_price == "" else "",
                                "" if out_of_stock == "" else "",
                                "" if quantity_available == "" else "",
                                "" if pd_url == ""else ""])
                            remaining -=1
                            failure_count +=1
                            update_status(output_text, f"❌ SKU {sku} Remaining {remaining}: Inactive")
            
            # ========== STEP 5: LOAD EXISTING SHEET (A TO C) ==========

            book = load_workbook(backup_file_path)
            sheet = book['Best Buy SKU List']  # Load the original sheet or wherever the data for A to C is stored.

            # Assuming you want to keep columns A to C, let's get that data into a DataFrame.
            df_existing = pd.read_excel(backup_file_path, sheet_name="Best Buy SKU List", usecols="A:C")

            # ========== STEP 6: CREATE RESULT DATAFRAME ==========

            # Reorganizing the columns for the final result (D to J)
            df_result = pd.DataFrame(product_list, columns=[
                "Status", "Buyable", "Discontinued", "Retail Price", 
                "Out Of Stock", "Quantity", "Product Link"
            ])

            # Combine the original data (A to C) with the result data (D to J)
            df_combined = pd.concat([df_existing, df_result], axis=1)

            # ========== STEP 7: REMOVE 'Result' SHEET IN BACKUP ==========

            # if 'Result' in book.sheetnames:
            #     del book['Result']

            # ========== STEP 8: SAVE THE COMBINED DATA BACK TO THE BACKUP FILE ==========

            if 'Result' in book.sheetnames:
                del book['Result']
                book.save(backup_file_path)  # Save after deleting to avoid error later

            # Write new sheet
            with pd.ExcelWriter(backup_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_combined.to_excel(writer, sheet_name="Result", index=False)

            # ========== STEP 9: FORMAT THE SHEET ==========

            # Reload to access newly written sheet
            book = load_workbook(backup_file_path)
            sheet_result = book["Result"]

            # Apply formatting...

            for row in sheet_result.iter_rows(min_row=1, max_row=sheet_result.max_row,
                                            min_col=1, max_col=sheet_result.max_column):
                for cell in row:
                    cell.font = Font(name="Calibri", size=12)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    cell.border = Border(
                        left=Side(border_style=None),
                        right=Side(border_style=None),
                        top=Side(border_style=None),
                        bottom=Side(border_style=None)
                    )

            # Format specific columns
            for col in ['A', 'C']:
                for cell in sheet_result[col]:
                    cell.number_format = '0'
                    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)

            for col in ['B', 'J']:
                for cell in sheet_result[col]:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

            for col_idx in range(1, 10):
                col_letter = get_column_letter(col_idx)
                max_length = 0

                # Loop through cells in the column to find the max length
                for cell in sheet_result[col_letter]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                # Set column width
                sheet_result.column_dimensions[col_letter].width = max_length + 2  # Adding a bit of padding

            # Freeze first row
            sheet_result.freeze_panes = sheet_result['A2']

            # Save workbook
            book.save(backup_file_path)

            # Summary
            end_time = datetime.datetime.now().time()
            # Calculate the time difference in seconds
            time_difference = datetime.datetime.combine(datetime.date.today(), end_time) - datetime.datetime.combine(datetime.date.today(), start_time)
            # Extract the time difference in seconds
            time_in_seconds = time_difference.total_seconds()
            
            update_status(output_text, f"\n🎯 Script complete! {time_in_seconds}")
            update_status(output_text, f"✅ Total Success: {success_count}")
            update_status(output_text, f"❌ Total Failures: {failure_count}")
            update_status(output_text, f"📄 Results saved to: {backup_file_path}")
            messagebox.showinfo("Alert", f"Script complete")

    threading.Thread(target=task).start()

def rgb_to_hex(r, g, b):
    return f"#{r:02x}{g:02x}{b:02x}"

def write_output(text):
    output_text.config(state=tk.NORMAL)
    output_text.insert(tk.END, text)
    output_text.see(tk.END)
    output_text.config(state=tk.DISABLED)

def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    position_top = int(screen_height / 2 - height / 2)
    position_right = int(screen_width / 2 - width / 2)
    window.geometry(f'{width}x{height}+{position_right}+{position_top}')

root = tk.Tk()
root.title("Memon v.1.15")
root.geometry("500x400")
root.resizable(False, False)
center_window(root, 500, 400)
# Use raw string (r"...") or double backslashes
# icon_path = os.path.join(os.path.dirname(__file__), "logo.ico")
# root.iconbitmap(icon_path)

button_font = ("Calibri", 12, "bold")
button_width = 15
button_height = 1

# Color Hex Codes from RGB
bestbuy_yellow = rgb_to_hex(255, 221, 0)      # #ffdd00
lowes_blue = rgb_to_hex(52, 130, 208)           # #0046ad
homedepot_orange = rgb_to_hex(245, 130, 32)   # #f58220
stop_red = rgb_to_hex(240, 65, 48)   # #f58220

# Define border colors (as RGB hex)
bestbuy_border = rgb_to_hex(200, 170, 0)       # Golden
lowes_border = rgb_to_hex(0, 35, 100)          # Deep blue
homedepot_border = rgb_to_hex(180, 80, 20)     # Deeper orange
stop_border = rgb_to_hex(255, 23, 0)     # Deeper orange

button_frame = tk.Frame(root)
button_frame.pack(pady=10)

# BestBuy
frame1 = tk.Frame(button_frame, bg=bestbuy_border, padx=2, pady=2)
frame1.pack(side=tk.LEFT, padx=5)

button1 = tk.Button(
    frame1, text="BestBuy SKU", command=run_bestbuy_script,
    bg=bestbuy_yellow, fg="black", font=button_font,
    width=button_width, height=button_height,
    relief="flat", borderwidth=0
)
button1.pack()

# Create a Text widget to display status
output_text = tk.Text(root, height=20, wrap=tk.WORD, state=tk.DISABLED)
output_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

# Add a scrollbar to the Text widget
scrollbar = tk.Scrollbar(root, command=output_text.yview)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
output_text.config(yscrollcommand=scrollbar.set)

# Bind window close event to stop the task
root.protocol("WM_DELETE_WINDOW", on_closing)
root.mainloop()